# Enhanced Surveillance Logging System
import os
from datetime import datetime
from enum import Enum
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class LogType(Enum):
    MOTION = "motion_detection"
    FACE_RECOGNITION = "face_recognition"
    ENTRY_EXIT = "entry_exit"
    RECORDING = "recording"
    RECTANGLE_MOTION = "rectangle_motion"
    SYSTEM = "system"

class SurveillanceLogger:
    def __init__(self):
        # Create logs directory if it doesn't exist
        self.log_dir = "logs"
        os.makedirs(self.log_dir, exist_ok=True)
        
        # Define Excel file path
        self.excel_file = f"{self.log_dir}/surveillance_logs.xlsx"
        
        # Define sheet names for different activities
        self.sheet_names = {
            LogType.MOTION: "Motion_Detection",
            LogType.FACE_RECOGNITION: "Face_Recognition",
            LogType.ENTRY_EXIT: "Entry_Exit",
            LogType.RECORDING: "Recording",
            LogType.RECTANGLE_MOTION: "Rectangle_Motion",
            LogType.SYSTEM: "System_Events"
        }
        
        # Initialize Excel file with sheets and headers
        self._initialize_excel_file()
    
    def _initialize_excel_file(self):
        """Initialize Excel file with sheets and headers"""
        headers = {
            LogType.MOTION: ["Timestamp", "Status", "Motion_Area", "Alert_Sent", "Image_Path", "Duration", "Sensitivity"],
            LogType.FACE_RECOGNITION: ["Timestamp", "Action", "Person_Name", "Person_ID", "Confidence", "Status", "Image_Count"],
            LogType.ENTRY_EXIT: ["Timestamp", "Direction", "Person_Count", "Image_Path", "Location", "Status"],
            LogType.RECORDING: ["Timestamp", "Action", "File_Path", "Duration", "Resolution", "Status"],
            LogType.RECTANGLE_MOTION: ["Timestamp", "Region", "Motion_Area", "Status", "Coordinates", "Sensitivity"],
            LogType.SYSTEM: ["Timestamp", "Module", "Action", "Status", "Details", "User"]
        }
        
        if not os.path.exists(self.excel_file):
            # Create new workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets with headers
            for log_type, sheet_name in self.sheet_names.items():
                ws = wb.create_sheet(title=sheet_name)
                
                # Add headers with styling
                header_row = headers[log_type]
                for col, header in enumerate(header_row, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
            
            wb.save(self.excel_file)
    
    
    def log_motion_detection(self, status, motion_area=0, alert_sent=False, image_path="", duration=0, sensitivity="medium"):
        """Log motion detection events"""
        self._write_log_entry(LogType.MOTION, [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            status,
            motion_area,
            "Yes" if alert_sent else "No",
            image_path,
            f"{duration:.2f}s" if duration > 0 else "N/A",
            sensitivity
        ])
    
    def log_face_recognition(self, action, person_name="Unknown", person_id="", confidence=0, status="", image_count=0):
        """Log face recognition events"""
        self._write_log_entry(LogType.FACE_RECOGNITION, [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            action,
            person_name,
            person_id,
            f"{confidence:.2f}%" if confidence > 0 else "N/A",
            status,
            image_count
        ])
    
    def log_entry_exit(self, direction, person_count=1, image_path="", location="main_entrance", status="detected"):
        """Log entry/exit events"""
        self._write_log_entry(LogType.ENTRY_EXIT, [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            direction,
            person_count,
            image_path,
            location,
            status
        ])
    
    def log_recording(self, action, file_path="", duration=0, resolution="640x480", status="in_progress"):
        """Log recording events"""
        self._write_log_entry(LogType.RECORDING, [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            action,
            file_path,
            f"{duration:.2f}s" if duration > 0 else "N/A",
            resolution,
            status
        ])
    
    def log_rectangle_motion(self, region, motion_area=0, status="no_motion", coordinates="", sensitivity="medium"):
        """Log rectangle motion detection events"""
        self._write_log_entry(LogType.RECTANGLE_MOTION, [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            region,
            motion_area,
            status,
            coordinates,
            sensitivity
        ])
    
    def log_system_event(self, module, action, status="info", details="", user="system"):
        """Log system events"""
        self._write_log_entry(LogType.SYSTEM, [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            module,
            action,
            status,
            details,
            user
        ])
    
    
    def _write_log_entry(self, log_type, data):
        """Write a log entry to the appropriate Excel sheet"""
        try:
            # Load existing workbook
            wb = load_workbook(self.excel_file)
            sheet_name = self.sheet_names[log_type]
            ws = wb[sheet_name]
            
            # Find next empty row
            next_row = ws.max_row + 1
            
            # Write data to row
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=next_row, column=col, value=value)
                
                # Apply alternating row colors for better readability
                if next_row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Color code based on status/severity
                if col == 2 and isinstance(value, str):  # Status column
                    if "error" in value.lower() or "failed" in value.lower():
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif "success" in value.lower() or "completed" in value.lower():
                        cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                    elif "warning" in value.lower() or "detected" in value.lower():
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            # Save workbook
            wb.save(self.excel_file)
            
        except Exception as e:
            print(f"Error writing to Excel file {self.excel_file}: {e}")
    
    def get_log_summary(self):
        """Get summary statistics for all Excel sheets"""
        summary = {}
        try:
            if os.path.exists(self.excel_file):
                wb = load_workbook(self.excel_file)
                for log_type, sheet_name in self.sheet_names.items():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        summary[log_type.value] = {
                            'total_entries': ws.max_row - 1,  # Subtract header
                            'last_modified': datetime.fromtimestamp(os.path.getmtime(self.excel_file)).strftime('%Y-%m-%d %H:%M:%S')
                        }
                    else:
                        summary[log_type.value] = {'total_entries': 0, 'status': 'Sheet not found'}
            else:
                for log_type in self.sheet_names.keys():
                    summary[log_type.value] = {'total_entries': 0, 'status': 'Excel file not found'}
        except Exception as e:
            for log_type in self.sheet_names.keys():
                summary[log_type.value] = {'error': str(e)}
        return summary

# Global logger instance
surveillance_logger = SurveillanceLogger()
