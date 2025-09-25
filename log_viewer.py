# Professional Log Viewer for Surveillance System
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
from datetime import datetime, timedelta
import subprocess
import platform
import pandas as pd
from openpyxl import load_workbook
from logger import LogType, surveillance_logger

class ProfessionalLogViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("Smart Surveillance - Activity Logs")
        self.root.geometry("1200x700")
        self.root.configure(bg='#f0f0f0')
        
        # Excel file path and sheet mapping
        self.excel_file = 'logs/surveillance_logs.xlsx'
        self.sheet_mapping = {
            'Motion Detection': 'Motion_Detection',
            'Face Recognition': 'Face_Recognition',
            'Entry/Exit': 'Entry_Exit',
            'Recording': 'Recording',
            'Rectangle Motion': 'Rectangle_Motion',
            'System Events': 'System_Events'
        }
        
        # Create main frame
        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create header
        self.create_header()
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill='both', expand=True, pady=(10, 0))
        
        # Create tabs for different log types
        self.create_log_tabs()
        
        # Create control panel
        self.create_control_panel()
        
        
        # Load initial data
        self.refresh_all_logs()
    
    def create_header(self):
        """Create professional header with summary"""
        header_frame = ttk.Frame(self.main_frame)
        header_frame.pack(fill='x', pady=(0, 10))
        
        # Title
        title_label = ttk.Label(header_frame, text="🔍 Smart Surveillance Activity Logs", 
                               font=('Arial', 16, 'bold'))
        title_label.pack(side='left')
        
        # Summary info
        self.summary_label = ttk.Label(header_frame, text="Loading...", 
                                      font=('Arial', 10))
        self.summary_label.pack(side='right')
        
        self.update_summary()
    
    def create_log_tabs(self):
        """Create tabs for different log types"""
        self.log_frames = {}
        self.log_trees = {}
        self.trees = {}
        
        log_configs = {
            'Motion Detection': {
                'type': LogType.MOTION,
                'columns': ['Timestamp', 'Status', 'Motion Area', 'Alert Sent', 'Image Path', 'Duration', 'Sensitivity'],
                'widths': [150, 100, 100, 80, 200, 80, 100]
            },
            'Face Recognition': {
                'type': LogType.FACE_RECOGNITION,
                'columns': ['Timestamp', 'Action', 'Person Name', 'Person ID', 'Confidence', 'Status', 'Image Count'],
                'widths': [150, 120, 120, 80, 80, 100, 100]
            },
            'Entry/Exit': {
                'type': LogType.ENTRY_EXIT,
                'columns': ['Timestamp', 'Direction', 'Person Count', 'Image Path', 'Location', 'Status'],
                'widths': [150, 100, 100, 200, 120, 100]
            },
            'Recording': {
                'type': LogType.RECORDING,
                'columns': ['Timestamp', 'Action', 'File Path', 'Duration', 'Resolution', 'Status'],
                'widths': [150, 100, 250, 80, 100, 100]
            },
            'Rectangle Motion': {
                'type': LogType.RECTANGLE_MOTION,
                'columns': ['Timestamp', 'Region', 'Motion Area', 'Status', 'Coordinates', 'Sensitivity'],
                'widths': [150, 100, 100, 100, 150, 100]
            },
            'System Events': {
                'type': LogType.SYSTEM,
                'columns': ['Timestamp', 'Module', 'Action', 'Status', 'Details', 'User'],
                'widths': [150, 120, 120, 100, 200, 100]
            }
        }
        
        for tab_name, config in log_configs.items():
            # Create tab frame
            tab_frame = ttk.Frame(self.notebook)
            self.notebook.add(tab_frame, text=tab_name)
            self.log_frames[config['type']] = tab_frame
            
            # Create treeview with scrollbars
            tree_frame = ttk.Frame(tab_frame)
            tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
            
            # Treeview
            tree = ttk.Treeview(tree_frame, columns=config['columns'], show='headings', height=20)
            
            # Configure columns
            for i, (col, width) in enumerate(zip(config['columns'], config['widths'])):
                tree.heading(col, text=col, command=lambda c=col: self.sort_column(tree, c, False))
                tree.column(col, width=width, minwidth=50)
            
            # Scrollbars
            v_scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
            h_scrollbar = ttk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
            tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
            
            # Pack scrollbars and tree
            tree.grid(row=0, column=0, sticky='nsew')
            v_scrollbar.grid(row=0, column=1, sticky='ns')
            h_scrollbar.grid(row=1, column=0, sticky='ew')
            
            tree_frame.grid_rowconfigure(0, weight=1)
            tree_frame.grid_columnconfigure(0, weight=1)
            
            self.log_trees[config['type']] = tree
            self.trees[tab_name] = tree
            
            # Add context menu
            self.add_context_menu(tree)
    
    def create_control_panel(self):
        """Create control panel with buttons and filters"""
        control_frame = ttk.Frame(self.main_frame)
        control_frame.pack(fill='x', pady=(10, 0))
        
        # Left side - buttons
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(side='left')
        
        ttk.Button(button_frame, text="🔄 Refresh All", 
                  command=self.refresh_all_logs).pack(side='left', padx=(0, 5))
        
        ttk.Button(control_frame, text="Export Current Tab", 
                   command=self.export_logs).pack(side='left', padx=5)
        
        ttk.Button(button_frame, text="🗑️ Clear Logs", 
                  command=self.clear_logs).pack(side='left', padx=5)
        
        ttk.Button(button_frame, text="📈 Statistics", 
                  command=self.show_statistics).pack(side='left', padx=5)
        
        # Right side - filters
        filter_frame = ttk.Frame(control_frame)
        filter_frame.pack(side='right')
        
        ttk.Label(filter_frame, text="Filter by date:").pack(side='left', padx=(0, 5))
        
        self.date_var = tk.StringVar(value="All")
        date_combo = ttk.Combobox(filter_frame, textvariable=self.date_var, 
                                 values=["All", "Today", "Yesterday", "Last 7 days", "Last 30 days"],
                                 width=12, state="readonly")
        date_combo.pack(side='left', padx=5)
        date_combo.bind('<<ComboboxSelected>>', self.apply_date_filter)
        
        ttk.Button(filter_frame, text="🔍 Search", 
                  command=self.show_search_dialog).pack(side='left', padx=(5, 0))
    
    def add_context_menu(self, tree):
        """Add right-click context menu to tree"""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="📋 Copy Row", command=lambda: self.copy_selected_row(tree))
        menu.add_command(label="📁 Open Image", command=lambda: self.open_image(tree))
        menu.add_separator()
        menu.add_command(label="🗑️ Delete Entry", command=lambda: self.delete_entry(tree))
        
        def show_context_menu(event):
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()
        
        tree.bind("<Button-3>", show_context_menu)
    
    def load_log_data(self, log_type):
        """Load log data from Excel sheet"""
        try:
            if not os.path.exists(self.excel_file):
                return []
            
            sheet_name = self.sheet_mapping.get(log_type)
            if not sheet_name:
                return []
            
            # Read Excel sheet using pandas
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            
            # Convert DataFrame to list of lists
            data = df.values.tolist()
            return data
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load {log_type} logs: {e}")
            return []
    
    def refresh_all_logs(self):
        """Refresh all log tabs"""
        self.apply_date_filter()
        self.update_summary()
    
    def update_summary(self):
        """Update summary information"""
        summary = surveillance_logger.get_log_summary()
        total_entries = sum(data['total_entries'] for data in summary.values())
        self.summary_label.config(text=f"Total Entries: {total_entries} | Last Updated: {datetime.now().strftime('%H:%M:%S')}")
    
    def sort_column(self, tree, col, reverse):
        """Sort tree column"""
        data = [(tree.set(child, col), child) for child in tree.get_children('')]
        data.sort(reverse=reverse)
        
        for index, (val, child) in enumerate(data):
            tree.move(child, '', index)
        
        tree.heading(col, command=lambda: self.sort_column(tree, col, not reverse))
    
    def load_log_data(self, log_type):
        """Load log data from Excel sheet"""
        try:
            if not os.path.exists(self.excel_file):
                return []
            
            sheet_name = self.sheet_mapping.get(log_type)
            if not sheet_name:
                return []
            
            # Read Excel sheet using pandas
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            
            # Convert DataFrame to list of lists
            data = df.values.tolist()
            return data
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load {log_type} logs: {e}")
            return []
    
    def apply_date_filter(self, event=None):
        """Apply date filter to current tab"""
        current_tab = self.notebook.select()
        tab_text = self.notebook.tab(current_tab, "text")
        
        # Find corresponding log type
        log_type = None
        for tab_name, sheet_name in self.sheet_mapping.items():
            if tab_name == tab_text:
                log_type = tab_name
                break
        
        if not log_type or log_type not in self.trees:
            return
        
        tree = self.trees[log_type]
        
        # Clear existing data
        for item in tree.get_children():
            tree.delete(item)
        
        # Get date filter
        date_filter = self.date_var.get()
        cutoff_date = None
        
        if date_filter == "Today":
            cutoff_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        elif date_filter == "Yesterday":
            cutoff_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
        elif date_filter == "Last 7 days":
            cutoff_date = datetime.now() - timedelta(days=7)
        elif date_filter == "Last 30 days":
            cutoff_date = datetime.now() - timedelta(days=30)
        # If "All" or any other value, cutoff_date remains None (show all data)
        
        # Load and filter data
        data = self.load_log_data(log_type)
        
        for row in data:
            if len(row) >= len(tree['columns']):
                if cutoff_date is None:  # Show all
                    tags = self.get_row_tags(log_type, row)
                    tree.insert('', 'end', values=row, tags=tags)
                else:
                    try:
                        row_date = datetime.strptime(str(row[0]), '%Y-%m-%d %H:%M:%S')
                        if row_date >= cutoff_date:
                            tags = self.get_row_tags(log_type, row)
                            tree.insert('', 'end', values=row, tags=tags)
                    except:
                        continue
    
    def get_row_tags(self, log_type, row):
        """Get tags for row coloring based on status"""
        if len(row) < 2:
            return ()
        
        status = str(row[1]).lower()
        if 'error' in status or 'failed' in status:
            return ('error',)
        elif 'success' in status or 'completed' in status:
            return ('success',)
        elif 'warning' in status or 'detected' in status:
            return ('warning',)
        else:
            return ('info',)
    
    def export_logs(self):
        """Export current filtered logs to CSV or Excel"""
        try:
            current_tab = self.notebook.tab(self.notebook.select(), "text")
            
            # Get current tree data
            tree = None
            for tab_name, tab_tree in self.trees.items():
                if tab_name == current_tab:
                    tree = tab_tree
                    break
            
            if not tree:
                messagebox.showwarning("Warning", "No data to export")
                return
            
            # Get file path
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")],
                title="Export Logs"
            )
            
            if filename:
                # Prepare data
                headers = [tree.heading(col)["text"] for col in tree["columns"]]
                data = []
                for item in tree.get_children():
                    values = tree.item(item)["values"]
                    data.append(values)
                
                # Create DataFrame
                df = pd.DataFrame(data, columns=headers)
                
                # Export based on file extension
                if filename.endswith('.xlsx'):
                    df.to_excel(filename, index=False, engine='openpyxl')
                else:
                    df.to_csv(filename, index=False)
                
                messagebox.showinfo("Success", f"Logs exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export logs: {e}")
    
    def clear_logs(self):
        """Clear all logs after confirmation"""
        result = messagebox.askyesno(
            "Confirm Clear", 
            "Are you sure you want to clear ALL surveillance logs?\n\nThis action cannot be undone!",
            icon="warning"
        )
        
        if result:
            try:
                if os.path.exists(self.excel_file):
                    # Load workbook and clear data from all sheets (keep headers)
                    wb = load_workbook(self.excel_file)
                    
                    for sheet_name in self.sheet_mapping.values():
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            # Delete all rows except header (row 1)
                            if ws.max_row > 1:
                                ws.delete_rows(2, ws.max_row)
                    
                    wb.save(self.excel_file)
                
                messagebox.showinfo("Success", "All logs have been cleared")
                self.refresh_all_logs()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to clear logs: {e}")
    
    def show_statistics(self):
        """Show statistics window"""
        stats_window = tk.Toplevel(self.root)
        stats_window.title("Log Statistics")
        stats_window.geometry("500x400")
        
        summary = surveillance_logger.get_log_summary()
        
        text_widget = tk.Text(stats_window, wrap='word', padx=10, pady=10)
        text_widget.pack(fill='both', expand=True)
        
        stats_text = "📊 SURVEILLANCE SYSTEM STATISTICS\n\n"
        for log_type, data in summary.items():
            stats_text += f"{log_type.replace('_', ' ').title()}:\n"
            if 'total_entries' in data:
                stats_text += f"  • Total Entries: {data['total_entries']}\n"
            if 'last_modified' in data:
                stats_text += f"  • Last Updated: {data['last_modified']}\n\n"
            elif 'error' in data:
                stats_text += f"  • Error: {data['error']}\n\n"
        
        text_widget.insert('1.0', stats_text)
        text_widget.config(state='disabled')
    
    def show_search_dialog(self):
        """Show search dialog"""
        search_window = tk.Toplevel(self.root)
        search_window.title("Search Logs")
        search_window.geometry("300x150")
        
        tk.Label(search_window, text="Search term:").pack(pady=10)
        search_var = tk.StringVar()
        tk.Entry(search_window, textvariable=search_var, width=30).pack(pady=5)
        
        def perform_search():
            # Implementation for search
            messagebox.showinfo("Search", f"Searching for: {search_var.get()}")
            search_window.destroy()
        
        tk.Button(search_window, text="Search", command=perform_search).pack(pady=10)
    
    def copy_selected_row(self, tree):
        """Copy selected row to clipboard"""
        selection = tree.selection()
        if selection:
            item = tree.item(selection[0])
            values = item['values']
            self.root.clipboard_clear()
            self.root.clipboard_append('\t'.join(str(v) for v in values))
    
    def open_image(self, tree):
        """Open image file if path exists in selected row"""
        selection = tree.selection()
        if selection:
            item = tree.item(selection[0])
            values = item['values']
            # Look for image path in values
            for value in values:
                if isinstance(value, str) and ('.jpg' in value or '.png' in value):
                    if os.path.exists(value):
                        os.startfile(value)
                        return
            messagebox.showinfo("No Image", "No image file found for this entry.")
    
    def delete_entry(self, tree):
        """Delete selected entry"""
        selection = tree.selection()
        if selection:
            if messagebox.askyesno("Confirm Delete", "Delete this log entry?"):
                tree.delete(selection[0])

def open_log_viewer():
    """Open the professional log viewer"""
    root = tk.Tk()
    app = ProfessionalLogViewer(root)
    root.mainloop()

if __name__ == "__main__":
    open_log_viewer()
